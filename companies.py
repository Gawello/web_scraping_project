import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def scrap_website(url):
    options = Options()
    options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)
    browser.get(url)

    # Przewijanie strony w dół
    last_height = browser.execute_script("return document.body.scrollHeight")
    while True:
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)  # Czekamy na załadowanie nowych elementów

        new_height = browser.execute_script("return document.body.scrollHeight")
        if new_height == last_height:  # Jeśli nie ma nowych elementów, zakończ przewijanie
            break
        last_height = new_height

    soup = BeautifulSoup(browser.page_source, 'html.parser')
    browser.quit()
    return soup


def extract_company_data(soup):
    # Wyodrębnianie informacji o firmie
    companies = soup.find_all('li', class_='card my-2')

    company_list = []

    for company in companies:
        company_name = company.find('a', class_='company-name').text.strip()
        address = company.find('div', class_='address').text.strip()
        phone = company.find('a', class_='icon-telephone').get('data-original-title').strip()
        email = company.find('a', class_='icon-envelope').get('data-company-email').strip()

        company_data = {
            'company_name': company_name,
            'address': address,
            'phone': phone,
            'email': email
        }
        company_list.append(company_data)

    return company_list


def save_to_excel(companies_data, filename):
    workbook = Workbook()
    sheet = workbook.active

    headers = ['Nazwa firmy', 'Adres', 'Telefon', 'Email']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    for row_num, company_data in enumerate(companies_data, 2):
        for col_num, key in enumerate(company_data, 1):
            sheet.cell(row=row_num, column=col_num).value = company_data[key]

    workbook.save(filename)


if __name__ == "__main__":
    response = scrap_website("https://panoramafirm.pl/palety_przemys%C5%82owe/polska")
    if not response:
        print("Connection error")
    else:
        company_data = extract_company_data(response)
        save_to_excel(company_data, 'firmy.xlsx')
